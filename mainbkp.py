from flask import Flask, render_template, redirect, url_for, flash, request, session, jsonify, Response

from flask_bootstrap import Bootstrap
from datetime import date
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import relationship
from flask_login import UserMixin, login_user, LoginManager, login_required, current_user, logout_user
from forms import CreateItemForm, RegisterForm, LoginForm
from functools import wraps
from flask import abort

from jinja2 import Environment, FileSystemLoader
from datetime import datetime

import openpyxl

from io import BytesIO

# Create Jinja2 environment
env = Environment(loader=FileSystemLoader('templates'))
env.globals.update(url_for=url_for)


def custom_date(date_str):
    """Custom Jinja2 filter to format date to dd/mm/yyyy"""
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')  # Convert date string to datetime object
    return date_obj.strftime('%d/%m/%Y')  # Format date to dd/mm/yyyy


# Add custom filter to Jinja2 environment
env.filters['custom_date'] = custom_date

app = Flask(__name__)
app.config['_permanent'] = False
app.config['SECRET_KEY'] = '8BYkEfBA6O6donzWlSihBXox7C0sKR6bb'
Bootstrap(app)

##CONNECT TO DB
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///stationary.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.session_protection = 'strong'
login_manager.login_view = 'login'
login_manager.init_app(app)


# APP_TOKEN = "wQyrp3mhYbAAAAAAAAAAHF3wr7tM2rmC0gEWN5pXolk"
# APP_TOKEN = "sl.BRflssQUvHnoAYZMmzSdLQscTgBF2L_vP4rr_QhF_CoJWBXW_VsIxk-P3Q5g3KovkvrYB4oYGDwwKLcQeJaGwMsDtNYxfHxM9QijIb45qznRRighvQQm2IF9Z9AbUA-AybRkw2UncnE"
# transferData = TransferData()


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# Create admin-only decorator

def admin_only(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # If id is not 1 then return abort with 403 error
        if current_user.id != 1:
            return abort(403)
        # Otherwise continue with the route function
        return f(*args, **kwargs)

    return decorated_function


##CONFIGURE TABLES


# CREATE USER TABLE
class User(UserMixin, db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(100))
    name = db.Column(db.String(100))

    user_requests = relationship("Request", back_populates="requested_by")


class Item(db.Model):
    __tablename__ = "items"
    id = db.Column(db.Integer, primary_key=True)
    item_name = db.Column(db.String(250), unique=True, nullable=False)
    item_code = db.Column(db.String(250), nullable=False)
    item_unit = db.Column(db.String(250), nullable=False)


class Request(db.Model):
    __tablename__ = "requests"
    id = db.Column(db.Integer, primary_key=True)
    req_date = db.Column(db.String(250), nullable=False)
    request_user_id = db.Column(db.Integer, db.ForeignKey('users.id'))

    requested_by = relationship("User", back_populates="user_requests")
    request_items = relationship("RequestItem", back_populates="parent_request")


class RequestItem(db.Model):
    __tablename__ = "request_items"
    id = db.Column(db.Integer, primary_key=True)
    req_id = db.Column(db.Integer, db.ForeignKey('requests.id'))
    item_id = db.Column(db.Integer, db.ForeignKey('items.id'))
    item_qty = db.Column(db.Integer, nullable=False)

    parent_request = relationship("Request", back_populates="request_items")


db.create_all()


@app.route('/', methods=["GET", "POST"])
def get_all():
    requests = Request.query.all()

    request_data = []
    for req in requests:
        request_items = req.request_items
        items = []
        for request_item in request_items:
            item_detail = Item.query.get(request_item.item_id)
            item = {
                "Name": item_detail.item_name,
                "Code": item_detail.item_code,
                "Quantity": request_item.item_qty,
                "Unit": item_detail.item_unit
            }
            items.append(item)

        req_data = {
            "Request ID": req.id,
            "Requested By": req.requested_by.name,
            "Date": req.req_date,
            "Items": items
        }
        request_data.append(req_data)
    template = env.get_template("index.html")  # Replace with the name of your template
    rendered_template = template.render(form=request_data,
                                        current_user=current_user)  # Replace with your actual form data
    return rendered_template  # Replace with your desired output (e.g. write to file, return as HTTP response, etc.)
    # return render_template("index.html", form=request_data, current_user=current_user)



@app.route('/new_request')
@login_required
def get_all_items():
    # page = request.args.get('page', 1, type=int)
    # all_items = Item.query.paginate(page, per_page=10)
    all_items = Item.query.all()
    return render_template("new_request.html", form=all_items, current_user=current_user)


# Add item to the table
@app.route('/add_item', methods=["GET", "POST"])
@login_required
def add():
    item_form = CreateItemForm()
    if item_form.validate_on_submit():
        # If user's email already exists
        if Item.query.filter_by(item_code=item_form.item_code.data).first():
            # Send flash messsage
            flash("This Item Code already exists, Use another one!")
            return redirect(url_for('add'))
        else:
            new_item = Item(
                item_code=item_form.item_code.data,
                item_name=item_form.item_name.data,
                item_unit=item_form.item_unit.data,

            )

            db.session.add(new_item)
            db.session.commit()
            return redirect(url_for('get_all_items'))

    return render_template("new_item.html", form=item_form, current_user=current_user)


# View the added item from the table
@app.route('/view_items', methods=['GET'])
@login_required
def view_items():
    all_items = Item.query.all()
    return render_template("view_list.html", form=all_items, current_user=current_user)


# Edit the item
@app.route("/edit-item/<int:item_id>", methods=["GET", "POST"])
@login_required
def edit_item(item_id):
    item = Item.query.get(item_id)
    edit_form = CreateItemForm(
        item_name=item.item_name,
        item_code=item.item_code,
        item_unit=item.item_unit

    )

    if edit_form.validate_on_submit():
        item.item_name = edit_form.item_name.data
        item.item_code = edit_form.item_code.data
        item.item_unit = edit_form.item_unit.data

        db.session.commit()
        return redirect(url_for("view_items"))

    return render_template("new_item.html", form=edit_form, is_edit=True, current_user=current_user)


@app.route("/delete/<int:item_id>")
@login_required
def delete_item(item_id):
    item_to_delete = Item.query.get(item_id)
    db.session.delete(item_to_delete)
    db.session.commit()

    return redirect(url_for('view_items'))


@app.route('/register', methods=["GET", "POST"])
def register():
    register_form = RegisterForm()
    if register_form.validate_on_submit():
        # If user's email already exists
        if User.query.filter_by(email=register_form.email.data).first():
            # Send flash messsage
            flash("You've already signed up with that email, log in instead!")
            # Redirect to /login route.
            # return redirect(url_for('login'))

        hash_and_salted_password = generate_password_hash(
            register_form.password.data,
            method='pbkdf2:sha256',
            salt_length=8
        )
        new_user = User(
            email=register_form.email.data,  # type: ignore
            name=register_form.name.data,  # type: ignore
            password=hash_and_salted_password,  # type: ignore
        )
        db.session.add(new_user)
        db.session.commit()
        login_user(new_user)
        return redirect(url_for("get_all_items"))
    return render_template("register.html", form=register_form, current_user=current_user)


@app.route('/login', methods=["GET", "POST"])
def login():
    login_form = LoginForm()
    if login_form.validate_on_submit():
        email = login_form.email.data
        password = login_form.password.data
        user = User.query.filter_by(email=email).first()
        # Email doesn't exist
        if not user:
            flash("That email does not exist, please try again.")
            return redirect(url_for('login'))
        # Password incorrect
        elif not check_password_hash(user.password, password):
            flash('Password incorrect, please try again.')
            return redirect(url_for('login'))
        else:
            login_user(user)
            session.permanent = False
            # print(f'user id = {session["_user_id"]}')
            return redirect(url_for('get_all_items'))
    return render_template("login.html", form=login_form, current_user=current_user)


# When the user click the Add to cart button, Create a cart varaible in Session and add the item to it with a default
# quantity of 1
@app.route('/add_to_cart', methods=["GET", "POST"])
def add_to_cart():
    if 'cart' not in session:
        session['cart'] = []
    if request.method == "POST":
        data = request.get_json()
        item_id = data["item_id"]
        # check if item_id already exists in the cart
        cart_item = next((item for item in session['cart'] if item["id"] == item_id), None)
        if cart_item:
            # item_id already exists, increase the quantity by 1
            pass
        else:
            # item_id does not exist, add it to the cart
            session['cart'].append({'id': item_id, 'quantity': 1})
        session.modified = True
        print(session['cart'])
    return redirect(url_for('get_all_items'))


@app.route('/remove_from_cart', methods=["GET", "POST"])
def remove_form_cart():
    if request.method == "POST":
        item_to_remove = request.get_json()
        id_to_remove = int(item_to_remove["del_Id"])  # convert to integer
        print(f"id_to_remove: {id_to_remove}")
        print(f"session['cart']: {session['cart']}")
        if 'cart' in session:
            cart_item = next((item for item in session['cart'] if item["id"] == str(id_to_remove)), None)

            print(f"cart_item: {cart_item}")
            if cart_item:
                session['cart'].remove(cart_item)
                session.modified = True
            print(session['cart'])
        return redirect(url_for('get_all_items'))


# Update the quantity in session when the user click the + or - Button
@app.route('/update_qty', methods=["GET", "POST"])
def update_qty():
    if request.method == "POST":
        data = request.get_json()

        # create a reference of existing cart
        cart_check = session['cart']
        # find the id of the item to update the quantity in session
        for index, value in enumerate(cart_check):
            if value["id"] == data["qty_id"]:
                cart_check[index]["quantity"] = data["upQty"]
                break  # exit the for loop
        session['cart'] = cart_check
        session.modified = True
        print(session['cart'])
    return redirect(url_for('get_all_items'))


@app.route('/create_request', methods=["GET", "POST"])
def create_request():
    if request.method == "POST":
        # create the request with date and the logged in user
        new_request = Request(req_date=date.today(), request_user_id=session['_user_id'])
        db.session.add(new_request)
        db.session.commit()
        # Get the id of the new request
        request_id = new_request.id
        # print(request_id)
        cart = session['cart']
        # save changes to the db with updated quantity and item id
        for index, item in enumerate(cart):
            new_request_item = RequestItem(req_id=request_id, item_id=cart[index]['id'],
                                           item_qty=cart[index]['quantity'])
            db.session.add(new_request_item)
            db.session.commit()
        session.pop('cart', None)
        return redirect(url_for('get_all'))
    return redirect(url_for('get_all'))

@app.route('/view_requests', methods=["GET", "POST"])
def get_all_requests():
    all_requests = Request.query.all()
    return render_template("view_requests_popup.html", form=all_requests, current_user=current_user)
    # requests = Request.query.options(joinedload(Request.request_items)).all()
    # result = []
    # for requests in requests:
    #     requested_items = []
    #     for request_item in requests.request_items:
    #         item_id = db.session.query(Item).filter(Item.id == request_item.item_id).first()
    #         item = {"id": request_item.item_id, "qty": request_item.item_qty, "name": item_id.item_name}
    #         requested_items.append(item)
    #     result.append({
    #         'id': requests.id,
    #         'req_date': requests.req_date,
    #         'requested_by': requests.requested_by.name,
    #         'requested_items': requested_items
    #     })
    # print(result)
    # return render_template("view_requests.html", form=result, current_user=current_user)


# @app.route("/view_requested_items/<int:req_id>", methods=["GET", "POST"])
# def view_requested_items(req_id):
#     all_requests = Request.query.all()
#     item_details = []
#     requested_items = RequestItem.query.filter_by(req_id=req_id).all()
#     for requested_item in requested_items:
#         item_id = db.session.query(Item).filter(Item.id == requested_item.item_id).first()
#         item = {"name": item_id.item_name, "item_code": item_id.item_code, "qty": requested_item.item_qty, "unit": item_id.item_unit}
#         item_details.append(item)
#     print(item_details)
#     print(len(requested_items))
#     return render_template("view_requests2.html", form=all_requests, requested_items=item_details, current_user=current_user)

@app.route("/view_requested_items", methods=["GET", "POST"])
def view_requested_items():
    if request.method == "POST":
        data = request.get_json()
        # print(data["req_id"])
        all_requests = Request.query.all()
        item_details = []
        requested_items = RequestItem.query.filter_by(req_id=data["req_id"]).all()
        for requested_item in requested_items:
            item_id = db.session.query(Item).filter(Item.id == requested_item.item_id).first()
            item = {"name": item_id.item_name, "item_code": item_id.item_code, "qty": requested_item.item_qty,
                    "unit": item_id.item_unit}
            item_details.append(item)
        # print(item_details)
        # print(len(requested_items))
        if request.headers['Content-Type'] == 'application/json':
            # print("json")
            return jsonify(item_details)
        return render_template("view_requests_popup.html", form=all_requests, requested_items=item_details,
                               current_user=current_user)


@app.route("/download/<int:req_id>", methods=["GET", "POST"])
def download_request(req_id):
    request_details = Request.query.filter_by(id=req_id).first()
    request_detail = []

    item_details = []
    if request_details:
        request_items = request_details.request_items
        req_data = {"Requested By": request_details.requested_by.name, "Date": request_details.req_date}
        request_detail.append(req_data)
        # print(request_detail)
        for request_item in request_items:
            item_id = request_item.item_id
            item_detail = Item.query.get(item_id)
            item = {"Name": item_detail.item_name, "Code": item_detail.item_code, "Quantity": request_item.item_qty,
                    "Unit": item_detail.item_unit}
            item_details.append(item)
    # print(item_details)
    # Load the Excel file
    book = openpyxl.load_workbook('ex1.xlsx')
    sheet = book.active
    # Write data to rows starting from row 11
    start_row = 11
    serial_number = 1
    for row, item in enumerate(item_details, start=start_row):
        sheet.cell(row=row, column=1).value = serial_number
        sheet.cell(row=row, column=2).value = item['Name'] + " " + item['Code']
        sheet.cell(row=row, column=3).value = item['Quantity']
        sheet.cell(row=row, column=5).value = item['Unit']
        serial_number += 1
    # Extract date value from the dictionary
    date_string = request_detail[0]['Date']

    # Convert date string to datetime object
    date_value = datetime.strptime(date_string, '%Y-%m-%d').date()

    # Change date format to DD/MM/YYYY
    date_formatted = date_value.strftime('%d/%m/%Y')

    # Write formatted date value to cell F8
    sheet['F8'].value = date_formatted
    # sheet['F8'] = request_detail[0]['Date']
    sheet['A29'] = "REQUESTED BY: " + request_detail[0]['Requested By']
    # sheet.cell(row=11, column=2).value = 2
    # book.save('write2cell.xlsx')

    # Save the updated dataframe to Excel
    # Create a file-like buffer to receive the output
    buffer = BytesIO()
    book.save(buffer)
    buffer.seek(0)

    # Send the Excel file as a response to the browser
    return Response(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=example.xlsx'})
    # df.to_excel(file_path, index=False)
    # return render_template("index.html", current_user=current_user)


@app.route('/logout')
def logout():
    logout_user()
    session.pop('cart', None)
    session.pop('_user_id', None)
    if '_fresh' in session:
        session.pop('_fresh')
    session.clear()
    return redirect(url_for('login'))


@app.route("/about")
def about():
    return render_template("about.html", current_user=current_user)


@app.route("/contact")
def contact():
    return render_template("contact.html", current_user=current_user)


@app.route("/view_all", methods=["GET", "POST"])
def view_all():
    requests = Request.query.all()
    request_data = []
    for req in requests:
        request_items = req.request_items
        items = []
        for request_item in request_items:
            item_detail = Item.query.get(request_item.item_id)
            item = {
                "Name": item_detail.item_name,
                "Code": item_detail.item_code,
                "Quantity": request_item.item_qty,
                "Unit": item_detail.item_unit
            }
            items.append(item)

        req_data = {
            "Request ID": req.id,
            "Requested By": req.requested_by.name,
            "Date": req.req_date,
            "Items": items
        }
        request_data.append(req_data)
    response = jsonify(request_data)
    print(response)
    # Set the Access-Control-Allow-Origin header to allow cross-origin requests
    response.headers.add('Access-Control-Allow-Origin', '*')
    return response


@app.route('/delete_request/<int:request_id>', methods=['DELETE'])
def delete_request(request_id):
    try:
        # Query the request by its ID
        request_to_delete = Request.query.get(request_id)
        if request_to_delete is not None:
            # Delete the associated request items first
            request_items_to_delete = RequestItem.query.filter_by(req_id=request_id).all()
            for request_item in request_items_to_delete:
                db.session.delete(request_item)

            # Delete the request itself
            db.session.delete(request_to_delete)
            db.session.commit()
            return jsonify({'message': 'Request and request items deleted successfully'})
        else:
            return jsonify({'error': 'Request not found'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete request: {}'.format(str(e))})


if __name__ == "__main__":
    app.run(debug=True)
